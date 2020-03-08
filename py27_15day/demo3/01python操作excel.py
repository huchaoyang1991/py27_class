"""
============================
Author:柠檬班-木森
Time:2020/3/7   11:02
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

import openpyxl

# 将指定的excel，加载为一个workbook对象
wb = openpyxl.load_workbook("cases.xlsx")

# 选中工作簿中的表单对象
sh = wb["musen"]

# 读取指定行列的格子对象
# cell = sh.cell(row=1, column=3)
# # 读取格子中的数据
# c12 = cell.value
# print(c12)

# 写入数据：
# sh.cell(row=3, column=2, value="python")
# # 将工作簿对象保存为文件
# wb.save("cases.xlsx")

# 获取表单中的最大行
# max_row = sh.max_row
# print(max_row)

# 获取表单中的最大列
# max_column = sh.max_column
# print(max_column)

# rows:按行获取表单中所有的格子对象，每行的内容，放在一个元组中
# res1 = list(sh.rows)
# print(res1)

# for item in res1:
#     for j in item:
#         print(j.value)

"""
[(<Cell 'musen'.A1>, <Cell 'musen'.B1>, <Cell 'musen'.C1>, <Cell 'musen'.D1>),
 (<Cell 'musen'.A2>, <Cell 'musen'.B2>, <Cell 'musen'.C2>, <Cell 'musen'.D2>),
 (<Cell 'musen'.A3>, <Cell 'musen'.B3>, <Cell 'musen'.C3>, <Cell 'musen'.D3>)
]

"""
# columns:按列获取表单中所有的格子对象，每列的内容，放在一个元组中

# res2 = list(sh.columns)
# print(res2)


# 需求：读取第三行的数据
# res = list(sh.rows)
# print(res[2])

# 需求：读取第二列到第四列的内容
# res1 = list(sh.columns)
# print(res1[1:4])










