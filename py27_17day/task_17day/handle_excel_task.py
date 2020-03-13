import openpyxl


class HandleExcelTask:
    """python 操作excel"""

    #     封装的需求：
    #     1、封装一个读数据的功能
    #     读数据需要用到什么？  excel文件路径，表单名
    #
    #
    # 2、封装一个写入数据的方法：
    #
    # 写数据需要用到哪些信息？ 文件，表单，行，列，写入内容
    def __init__(self, file_name, sheet_name):
        """

        :param filename:excel文件路径
        :param sheetname:表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self):
        """读取excel数据"""
        work = openpyxl.load_workbook(self.file_name)
        sh = work[self.sheet_name]
        rows_data = list(sh.rows)
        cases_data = []
        title = []
        for i in rows_data[0]:
            title.append(i.value)
        for item in rows_data[1:]:
            values = []
            for i in item:
                values.append(i.value)
            case = dict(zip(title, values))
            cases_data.append(case)
        return cases_data

    def write_data(self, row, column, value):
        """写入excel数据"""
        work = openpyxl.load_workbook(self.file_name)
        sh = work[self.sheet_name]
        sh.cell(row=row, column=column, value=value)
        work.save(self.file_name)
