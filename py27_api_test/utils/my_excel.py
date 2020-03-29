import openpyxl


class MyExcel:
    def __init__(self, filename, sheetname):
        """
        初始化对象属性
        :param filename: excel文件路径
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """
        读取excel数据
        :return: 返回列表（字典）
        """
        cases_data = []  # 用例数据
        title = []  # 表头信息
        work = openpyxl.load_workbook(self.filename)
        sh = work[self.sheetname]

        rows_data = list(sh.rows)
        for i in rows_data[0]:
            title.append(i.value)

        for item in rows_data[1:]:
            values = []
            for i in item:
                values.append(i.value)
            cases_data.append(dict(zip(title, values)))
        return cases_data

    def write_data(self, row, column, value):
        work = openpyxl.load_workbook(self.filename)
        sh = work[self.sheetname]
        sh.cell(row, column, value)
        work.save(self.filename)
